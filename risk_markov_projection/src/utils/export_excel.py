from __future__ import annotations

from pathlib import Path
from typing import Iterable, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import config
from src.utils.metrics import aggregate_indicator_by_mob, mae, wape

HEADER_COLOR = "FFC000"
HEADER_COL_GRAY = "D9D9D9"


def _auto_width(ws) -> None:
    for col in ws.columns:
        max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 60)


def export_projection_excel(
    projection_df: pd.DataFrame,
    output_path: str | Path,
    actual_del30p_by_mob: Optional[pd.Series] = None,
) -> Path:
    """
    Export projection results to Excel with per-segment sheets and a summary.

    Parameters
    ----------
    projection_df : pd.DataFrame
        Output from the projection pipeline containing EAD_/DIST_/DEL_/audit columns.
    output_path : str | Path
        Path to the Excel file to create.
    actual_del30p_by_mob : pd.Series, optional
        Optional series indexed by MOB with observed DEL_30P_ON_EAD0 to compare against.
    """
    if projection_df.empty:
        raise ValueError("projection_df is empty; nothing to export.")

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    mob_col = config.SCHEMA["mob_col"]
    segment_cols = list(config.SCHEMA["segment_cols"])

    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = "Summary"
    summary_ws.sheet_view.showGridLines = False

    # Summary header
    summary_headers = [
        *segment_cols,
        "FallbackRate",
        "MeanCalibration",
        "Rows",
        "MAE_DEL30P",
        "WAPE_DEL30P",
    ]
    summary_ws.append(summary_headers)
    for cell in summary_ws[1]:
        cell.fill = PatternFill("solid", fgColor=HEADER_COL_GRAY)
        cell.font = Font(bold=True)

    # Per-segment sheets
    for _, seg_row in projection_df[segment_cols].drop_duplicates().iterrows():
        mask = (projection_df[segment_cols] == seg_row.values).all(axis=1)
        seg_df = projection_df.loc[mask].copy()
        seg_df = seg_df.sort_values(mob_col)

        # Align actuals
        actual_series = actual_del30p_by_mob.reindex(seg_df[mob_col]).reset_index(drop=True) if actual_del30p_by_mob is not None else None
        seg_df = seg_df.reset_index(drop=True)
        if actual_series is not None:
            seg_df["ACTUAL_DEL30P_ON_EAD0"] = actual_series.values
            seg_df["ABS_ERR"] = (seg_df["ACTUAL_DEL30P_ON_EAD0"] - seg_df["DEL_30P_ON_EAD0"]).abs()
            seg_df["REL_ERR"] = seg_df["ABS_ERR"] / seg_df["ACTUAL_DEL30P_ON_EAD0"].abs().replace({0: pd.NA})

        # Columns to keep
        base_cols = [
            mob_col,
            "EAD0",
            "DEL_30P_ON_EAD0",
            "DEL_60P_ON_EAD0",
            "DEL_90P_ON_EAD0",
        ]
        audit_cols = ["matrix_source", "mob_used", "n_obs_used", "ead_sum_used", "calibration_factor"]
        extra_cols = [c for c in seg_df.columns if c.startswith("DIST_") or c.startswith("EAD_")]
        ordered_cols = base_cols.copy()
        if "ACTUAL_DEL30P_ON_EAD0" in seg_df:
            ordered_cols += ["ACTUAL_DEL30P_ON_EAD0", "ABS_ERR", "REL_ERR"]
        ordered_cols += audit_cols
        ordered_cols += extra_cols
        ordered_cols = [c for c in ordered_cols if c in seg_df.columns]

        ws_name = "-".join(str(seg_row[c]) for c in segment_cols)[:31] or "Segment"
        ws = wb.create_sheet(title=ws_name)
        ws.sheet_view.showGridLines = False

        ws.append(ordered_cols)
        for cell in ws[1]:
            cell.fill = PatternFill("solid", fgColor=HEADER_COLOR)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        pct_cols = {"DEL_30P_ON_EAD0", "DEL_60P_ON_EAD0", "DEL_90P_ON_EAD0", "ACTUAL_DEL30P_ON_EAD0", "ABS_ERR", "REL_ERR"}
        for _, row in seg_df[ordered_cols].iterrows():
            cleaned = []
            for val in row.tolist():
                if pd.isna(val):
                    cleaned.append(None)
                else:
                    cleaned.append(val)
            ws.append(cleaned)

        # Format percentage-like columns
        for col_idx, col_name in enumerate(ordered_cols, start=1):
            if col_name in pct_cols:
                for r in range(2, ws.max_row + 1):
                    ws.cell(row=r, column=col_idx).number_format = "0.00%"

        ws.freeze_panes = "A2"
        _auto_width(ws)

        # Summary row
        fallback_rate = float((seg_df["matrix_source"] != "segment_mob").mean()) if "matrix_source" in seg_df else float("nan")
        mean_calib = float(seg_df["calibration_factor"].mean()) if "calibration_factor" in seg_df else float("nan")

        mae_val = wape_val = float("nan")
        if actual_series is not None and not seg_df["ACTUAL_DEL30P_ON_EAD0"].isna().all():
            mae_val = mae(seg_df["ACTUAL_DEL30P_ON_EAD0"], seg_df["DEL_30P_ON_EAD0"])
            wape_val = wape(seg_df["ACTUAL_DEL30P_ON_EAD0"], seg_df["DEL_30P_ON_EAD0"])

        summary_ws.append([
            *[seg_row[c] for c in segment_cols],
            fallback_rate,
            mean_calib,
            len(seg_df),
            mae_val,
            wape_val,
        ])

    _auto_width(summary_ws)

    wb.save(output_path)
    return output_path
