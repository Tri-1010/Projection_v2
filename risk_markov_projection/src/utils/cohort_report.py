from __future__ import annotations

from pathlib import Path
from typing import Dict, Iterable, List, Optional

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

import config
from src.data.schema import DataSchema, default_schema


def _del30_by_cohort_actual(raw_df: pd.DataFrame, schema: DataSchema, buckets: List[str], max_mob: int) -> pd.DataFrame:
    mob_col, ead_col, state_col = schema.mob_col, schema.ead_col, schema.state_col
    cohort_col = schema.cohort_col
    if not cohort_col or cohort_col not in raw_df.columns:
        raise ValueError("cohort_col is not set or missing in raw data.")

    df = raw_df.copy()
    df[cohort_col] = pd.to_datetime(df[cohort_col]).dt.to_period("M").dt.to_timestamp()

    ead0 = df[df[mob_col] == 0].groupby(cohort_col)[ead_col].sum()
    cohorts = ead0.index
    mob_values = list(range(max_mob + 1))

    actual_cols: Dict[str, pd.Series] = {}
    for mob in mob_values:
        g = df[df[mob_col] == mob]
        if g.empty:
            continue
        numer = g[g[state_col].isin(buckets)].groupby(cohort_col)[ead_col].sum()
        ratio = numer.reindex(cohorts).divide(ead0, fill_value=0).fillna(0)
        actual_cols[f"MOB{mob}_ACTUAL"] = ratio

    if not actual_cols:
        return pd.DataFrame()

    actual_df = pd.DataFrame(actual_cols)
    actual_df.insert(0, "Cohort", actual_df.index)
    return actual_df.reset_index(drop=True)


def _del30_by_cohort_forecast(raw_df: pd.DataFrame, projection_df: pd.DataFrame, schema: DataSchema, max_mob: int) -> pd.DataFrame:
    cohort_col = schema.cohort_col
    if not cohort_col or cohort_col not in raw_df.columns:
        raise ValueError("cohort_col is not set or missing in raw data.")

    segment_cols = list(schema.segment_cols)
    mob_col = schema.mob_col
    ead_col = schema.ead_col

    df = raw_df.copy()
    df[cohort_col] = pd.to_datetime(df[cohort_col]).dt.to_period("M").dt.to_timestamp()

    # EAD0 weights by cohort and segment
    ead0_seg = (
        df[df[mob_col] == 0]
        .groupby([cohort_col] + segment_cols)[ead_col]
        .sum()
    )
    if ead0_seg.empty:
        return pd.DataFrame()
    ead0_total = ead0_seg.groupby(level=0).sum()

    seg_del = (
        projection_df
        .set_index(segment_cols + [mob_col])["DEL_30P_ON_EAD0"]
        .to_frame("del30")
        .reset_index()
    )
    mob_values = list(range(max_mob + 1))

    records: Dict[str, Dict[pd.Timestamp, float]] = {}
    for cohort, cohort_seg in ead0_seg.groupby(level=0):
        total = ead0_total.loc[cohort]
        weight_df = (cohort_seg / total).reset_index()
        weight_df.columns = [cohort_col] + segment_cols + ["weight"]
        for mob in mob_values:
            merged = weight_df.merge(
                seg_del[seg_del[mob_col] == mob],
                on=segment_cols,
                how="left",
            )
            merged["del30"] = merged["del30"].fillna(0)
            merged["contrib"] = merged["weight"] * merged["del30"]
            val = merged["contrib"].sum()
            records.setdefault(f"MOB{mob}_FORECAST", {})[cohort] = val

    if not records:
        return pd.DataFrame()

    forecast_df = pd.DataFrame(records)
    forecast_df.insert(0, "Cohort", forecast_df.index)
    return forecast_df.reset_index(drop=True)


def build_cohort_del30_report(
    raw_df: pd.DataFrame,
    projection_df: pd.DataFrame,
    schema: Optional[DataSchema] = None,
    state_order: Optional[List[str]] = None,
    buckets_30p: Optional[List[str]] = None,
    max_mob: Optional[int] = None,
) -> pd.DataFrame:
    """Create a Cohort x MOB table for DEL30% over EAD0, with actual and forecast columns."""
    schema = schema or default_schema()
    state_order = state_order or config.STATE_ORDER
    buckets_30p = buckets_30p or config.BUCKETS_30P
    max_mob = max_mob if max_mob is not None else config.MAX_MOB

    actual_df = _del30_by_cohort_actual(raw_df, schema, buckets_30p, max_mob)
    forecast_df = _del30_by_cohort_forecast(raw_df, projection_df, schema, max_mob)

    if actual_df.empty and forecast_df.empty:
        return pd.DataFrame()

    if actual_df.empty:
        result = forecast_df
    elif forecast_df.empty:
        result = actual_df
    else:
        result = pd.merge(actual_df, forecast_df, on="Cohort", how="outer")

    result = result.sort_values("Cohort").reset_index(drop=True)
    return result


def export_cohort_del30_report(
    raw_df: pd.DataFrame,
    projection_df: pd.DataFrame,
    output_path: str | Path,
    schema: Optional[DataSchema] = None,
    state_order: Optional[List[str]] = None,
    buckets_30p: Optional[List[str]] = None,
    max_mob: Optional[int] = None,
) -> Path:
    report_df = build_cohort_del30_report(
        raw_df,
        projection_df,
        schema=schema,
        state_order=state_order,
        buckets_30p=buckets_30p,
        max_mob=max_mob,
    )
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    report_df.to_csv(output_path, index=False)
    return output_path


def _auto_width(ws) -> None:
    for col in ws.columns:
        max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 60)


def _extract_mobs(df: pd.DataFrame, suffix: str) -> List[int]:
    mobs = []
    for col in df.columns:
        if col.startswith("MOB") and col.endswith(suffix):
            try:
                mobs.append(int(col[3:-len(suffix)]))
            except ValueError:
                continue
    return mobs


def export_cohort_del30_excel_combined(
    raw_df: pd.DataFrame,
    projection_df: pd.DataFrame,
    output_path: str | Path,
    schema: Optional[DataSchema] = None,
    state_order: Optional[List[str]] = None,
    buckets_30p: Optional[List[str]] = None,
    max_mob: Optional[int] = None,
) -> Path:
    """
    Export Cohort x MOB table with a single column per MOB.
    - If actual exists for a MOB, use actual.
    - Else use forecast and style in red bold.
    All MOB cells are formatted as percentages.
    """
    schema = schema or default_schema()
    state_order = state_order or config.STATE_ORDER
    buckets_30p = buckets_30p or config.BUCKETS_30P
    max_mob = max_mob if max_mob is not None else config.MAX_MOB

    actual_df = _del30_by_cohort_actual(raw_df, schema, buckets_30p, max_mob)
    forecast_df = _del30_by_cohort_forecast(raw_df, projection_df, schema, max_mob)

    if actual_df.empty and forecast_df.empty:
        raise ValueError("No cohort data available for export.")

    actual_df = actual_df.set_index("Cohort") if not actual_df.empty else pd.DataFrame()
    forecast_df = forecast_df.set_index("Cohort") if not forecast_df.empty else pd.DataFrame()

    mobs = sorted(set(_extract_mobs(actual_df, "_ACTUAL") + _extract_mobs(forecast_df, "_FORECAST")))
    cohorts = sorted(set(actual_df.index.tolist()) | set(forecast_df.index.tolist()))

    wb = Workbook()
    ws = wb.active
    ws.title = "Cohort DEL30"
    ws.sheet_view.showGridLines = False

    header = ["Cohort"] + [f"MOB{m}" for m in mobs]
    ws.append(header)
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor="D9D9D9")
        cell.font = Font(bold=True)

    forecast_font = Font(color="FF0000", bold=True)

    for cohort in cohorts:
        row_vals = [cohort]
        row_fonts = [None]
        for mob in mobs:
            act_col = f"MOB{mob}_ACTUAL"
            fore_col = f"MOB{mob}_FORECAST"
            act_val = actual_df.at[cohort, act_col] if not actual_df.empty and act_col in actual_df.columns and cohort in actual_df.index else np.nan
            fore_val = forecast_df.at[cohort, fore_col] if not forecast_df.empty and fore_col in forecast_df.columns and cohort in forecast_df.index else np.nan
            if pd.notna(act_val) and (float(act_val) != 0 or pd.isna(fore_val) or float(fore_val) == 0):
                # Prefer actual when it is non-zero, or when forecast is missing/zero
                val = float(act_val)
                font = None
            elif pd.notna(fore_val):
                # Use forecast (highlight) when actual is missing or zero but forecast has signal
                val = float(fore_val)
                font = forecast_font
            else:
                val = None
                font = None
            row_vals.append(val)
            row_fonts.append(font)

        ws.append(row_vals)
        row_idx = ws.max_row
        for col_idx, font in enumerate(row_fonts, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if col_idx > 1:
                cell.number_format = "0.00%"
            if font:
                cell.font = font

    _auto_width(ws)
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def export_cohort_del30_excel_split(
    raw_df: pd.DataFrame,
    projection_df: pd.DataFrame,
    output_path: str | Path,
    schema: Optional[DataSchema] = None,
    state_order: Optional[List[str]] = None,
    buckets_30p: Optional[List[str]] = None,
    max_mob: Optional[int] = None,
) -> Path:
    """
    Export Cohort x MOB table with separate ACTUAL and FORECAST columns per MOB.
    Forecast cells are styled in red bold; all MOB cells formatted as %.
    """
    schema = schema or default_schema()
    state_order = state_order or config.STATE_ORDER
    buckets_30p = buckets_30p or config.BUCKETS_30P
    max_mob = max_mob if max_mob is not None else config.MAX_MOB

    actual_df = _del30_by_cohort_actual(raw_df, schema, buckets_30p, max_mob)
    forecast_df = _del30_by_cohort_forecast(raw_df, projection_df, schema, max_mob)

    if actual_df.empty and forecast_df.empty:
        raise ValueError("No cohort data available for export.")

    actual_df = actual_df.set_index("Cohort") if not actual_df.empty else pd.DataFrame()
    forecast_df = forecast_df.set_index("Cohort") if not forecast_df.empty else pd.DataFrame()

    mobs = sorted(set(_extract_mobs(actual_df, "_ACTUAL") + _extract_mobs(forecast_df, "_FORECAST")))
    cohorts = sorted(set(actual_df.index.tolist()) | set(forecast_df.index.tolist()))

    wb = Workbook()
    ws = wb.active
    ws.title = "Cohort DEL30 split"
    ws.sheet_view.showGridLines = False

    header = ["Cohort"] + [f"MOB{m}_ACTUAL" for m in mobs] + [f"MOB{m}_FORECAST" for m in mobs]
    ws.append(header)
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor="D9D9D9")
        cell.font = Font(bold=True)

    forecast_font = Font(color="FF0000", bold=True)

    for cohort in cohorts:
        row_vals = [cohort]
        row_fonts = [None]
        for mob in mobs:
            act_col = f"MOB{mob}_ACTUAL"
            val = None
            if not actual_df.empty and act_col in actual_df.columns and cohort in actual_df.index:
                val = actual_df.at[cohort, act_col]
            row_vals.append(val if pd.notna(val) else None)
            row_fonts.append(None)
        for mob in mobs:
            fore_col = f"MOB{mob}_FORECAST"
            val = None
            if not forecast_df.empty and fore_col in forecast_df.columns and cohort in forecast_df.index:
                val = forecast_df.at[cohort, fore_col]
            row_vals.append(val if pd.notna(val) else None)
            row_fonts.append(forecast_font if pd.notna(val) else None)

        ws.append(row_vals)
        row_idx = ws.max_row
        for col_idx, font in enumerate(row_fonts, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if col_idx > 1:
                cell.number_format = "0.00%"
            if font:
                cell.font = font

    _auto_width(ws)
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path
